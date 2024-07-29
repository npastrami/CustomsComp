import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import numpy as np
import pandas as pd
import re
from io import BytesIO
from itemcodeoffsets import keyword_to_offset_dict

def process_FOF(workbook, fof_sheets):   
        mappings = {
            "shippername": 2,
            "shipperaddress": 3,
            "shipperphone": 4,
            "invoicenum": 5,
            "countrycode": 6,
            "date": 7,
            "mawbnum": 8,
            "hawbnum": 9,
            "airlineandflightnum": 10,
            "freightforwarder": 11,
            "rucnum": 12,
            "daenum": 13,
            "incoterm": 14,
            "consigneename": 15,
            "consigneeaddress": 16,
            "consigneecityc": 17,
            "consigneephone": 18,
            "consigneecontact": 19,
            "consigneepostalcode": 20,
            "fixedprice": 21,
            "consignment": 22,
            "piecestype1": 23,
            "piecestype2": 24,
            "piecestype3": 25,
            "piecestype4": 26,
            "piecestype5": 27,
            "totalpices1": 28,
            "totalpices2": 29,
            "totalpices3": 30,
            "totalpices4": 31,
            "totalpices5": 32,
            "eqfullboxes1": 33,
            "eqfullboxes2": 34,
            "eqfullboxes3": 35,
            "eqfullboxes4": 36,
            "eqfullboxes5": 37,
            "product1": 38,
            "product2": 39,
            "product3": 40,
            "product4": 41,
            "product5": 42,
            "hits#1": 43,
            "hits#2": 44,
            "hits#3": 45,
            "hits#4": 46,
            "hits#5": 47,
            "nandina1": 48,
            "nandina2": 49,
            "nandina3": 50,
            "nandina4": 51,
            "nandina5": 52,
            "totalunits1": 53,
            "totalunits2": 54,
            "totalunits3": 55,
            "totalunits4": 56,
            "totalunits5": 57,
            "stemsbunch1": 58,
            "stemsbunch2": 59,
            "stemsbunch3": 60,
            "stemsbunch4": 61,
            "stemsbunch5": 62,
            "unitprice1": 63,
            "unitprice2": 64,
            "unitprice3": 65,
            "unitprice4": 66,
            "unitprice5": 67,
            "totalvalue1": 68,
            "totalvalue2": 69,
            "totalvalue3": 70,
            "totalvalue4": 71,
            "totalvalue5": 72,
            "samples": 73,
            "billto": 74,
            "shippercontact": 75,
            "shipperfax": 76,
            "consigneefax": 77,
        }
        
        target_worksheet = workbook["Sheet1"]
        starting_row = 5

        for sheet_index, fof_sheet in enumerate(fof_sheets, start=starting_row):
            # Extracting the sheet name and removing the 'FOF_' prefix
            sheet_title = fof_sheet.title.replace("FOF_", "")
            row_number = sheet_index
            target_worksheet[f'A{row_number}'] = sheet_title
            
            #list of lists to store items with footnotes
            items_with_footnotes = []
            items_with_target_col = []

            for row in fof_sheet.iter_rows(min_row=1):
                keyword_cell = row[0].value
                if keyword_cell:
                    # Remove brackets from keyword_cell for comparison
                    keyword_for_comparison = re.sub(r'[\[\]]', '', keyword_cell).strip()
                    if keyword_for_comparison in mappings:
                        # Fetch item code associated with keyword instance
                        item_code = row[1].value

                        if item_code and '*' in item_code:
                            item_code = item_code.replace('*', '')
                            items_with_footnotes.append([keyword_cell, item_code, int(sheet_index)])  # use original keyword_cell with brackets for later reference

                        if keyword_for_comparison in keyword_to_offset_dict:
                            offset_dict = keyword_to_offset_dict[keyword_for_comparison]

                            if item_code in offset_dict:
                                offset = offset_dict[item_code]
                                target_col = int(mappings[keyword_for_comparison]) + offset + 1  # Convert to int
                                print(f'Target col for {keyword_cell} with item code {item_code} is {target_col}')
                            else:
                                invalid_item_codes = {keyword_cell: item_code}
                                print(f'This Keyword-Item Code pairing is invalid: {invalid_item_codes}')
                                continue
                        else:
                            # For keywords without offsets
                            target_col = int(mappings[keyword_for_comparison]) + 1  # Convert to int
                        print(f'Final target col: {target_col}, for {keyword_cell} with item code {item_code}')
                        # Append target col to items_with_target_col matched with base keyword and item code
                        for item in items_with_footnotes:
                            if item[0] == keyword_cell and item[1] == item_code:
                                items_with_target_col.append(item + [target_col])  # Create a new list with target_col

                        amount_cell = row[2].value
                        target_worksheet.cell(row=row_number, column=target_col, value=amount_cell)
            print(f'All items with footnotes: {items_with_footnotes}')
            
            for keyword_for_comparison, item_code, sheet_index, target_col in items_with_target_col:
                # Ensure target_row and sheet_index are integers before calling the cell method
                target_col = int(target_col) if isinstance(target_col, str) else target_col
                sheet_index = int(sheet_index) if isinstance(sheet_index, str) else sheet_index
                cell_to_highlight = target_worksheet.cell(row=row_number, column=target_col)
                cell_to_highlight.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        
        for row in range(1, target_worksheet.max_row + 1):
            # row_letter = get_column_letter(row)
            target_worksheet.row_dimensions[row].height = 25  # Approximate conversion
                
        # Make Totals Column for Rows 5-300 in Column A
        for col_num in range(5, 300):  # End range is exclusive, so 300 to include column 299
            sum_value = 0  # Initialize sum for the current column
            for row_num in range(1, target_worksheet.max_row + 1):  # Start from row 1
                cell = target_worksheet.cell(row=row_num, column=col_num)
                cell_value = cell.value
                # Check if the cell contains a string with parentheses or a negative sign
                if cell_value is not None:
                    # Remove commas, parentheses, and negative signs for numeric values
                    if isinstance(cell_value, str):
                        cell_value_cleaned = cell_value.replace(',', '').replace('(', '').replace(')', '').replace('-', '')
                        if cell_value_cleaned.isdigit():
                            numeric_value = int(cell_value_cleaned)
                            if '(' in cell_value or '-' in cell_value:
                                sum_value -= numeric_value
                            else:
                                sum_value += numeric_value

            # Update Row 1 with the calculated sum for the current column, converting sum to string, make 0 totals blank
            if sum_value != 0:
                target_worksheet.cell(row=1, column=col_num, value=str(sum_value))  # Row 1 is the 1st row
            else:
                target_worksheet.cell(row=1, column=col_num, value="")

        # Set text wrap for column A
        for row in range(1, target_worksheet.max_row + 1):
            cell = target_worksheet.cell(row=row, column=1)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Rename the worksheet
        target_worksheet.title = "FOF_Data"