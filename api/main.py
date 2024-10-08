from quart import Quart, request, jsonify, Response, send_file
from quart_cors import cors
from werkzeug.utils import secure_filename
from uploader import Uploader
from extractor import Extractor
from azure.storage.blob.aio import BlobServiceClient
import azure_credentials 
import aiofiles
from io import BytesIO
from copyfunc import copy_worksheet
from database import Database
from FOFexport import process_FOF
from sorter import Sorter
from form_mapping_utils import upload_bucket_mapping
from table_builder import TableBuilder
import xml.etree.ElementTree as ET
import re
import asyncio
import openpyxl
import io

app = Quart(__name__)
app = cors(app)

@app.route('/process_doc', methods=['POST'])
async def process_doc():
    form = await request.form
    client_id = form['clientID']
    version_id = form['versionID']
    form_types = form.getlist('formTypes[]')
    uploaded_files = (await request.files).getlist('files[]')

    tasks = []

    for uploaded_file, form_type in zip(uploaded_files, form_types):
        tasks.append(handle_file(client_id, uploaded_file, form_type, version_id))

    responses = await asyncio.gather(*tasks)

    return jsonify(responses)

async def handle_file(client_id, uploaded_file, form_type, version_id):
    filename = secure_filename(uploaded_file.filename)
    blob_url = await upload_file(client_id, uploaded_file, form_type, version_id)

    if not blob_url:
        return {"status": "Error", "error": "Upload failed"}

    if form_type != 'None':
        xml_str = await extract_data(client_id, filename, upload_bucket_mapping[form_type], form_type, version_id)
        if xml_str:
            return {"status": "Extract Completed", "xml": xml_str}
        else:
            return {"status": "Empty Extraction"}
    else:
        return {"status": "Upload Completed", "uploaded_file": blob_url}

async def upload_file(client_id, uploaded_file, form_type, version_id):
    bucket_name = upload_bucket_mapping.get(form_type, 'unsorted')
    uploader = Uploader(bucket_name)
    blob_url = await uploader.upload(client_id, uploaded_file)
    if blob_url:
        database = Database(client_id, blob_url)
        await database.post2postgres_upload(client_id, blob_url, 'uploaded', form_type, bucket_name, version_id)
    return blob_url

async def extract_data(client_id, filename, bucket_name, form_type, version_id):
    sanitized_blob_name = sanitize_blob_name(filename)
    extractor = Extractor(bucket_name)
    extracted_values, blob_sas_url = await extractor.extract(client_id, sanitized_blob_name, form_type)
    await extractor.update_database(client_id, blob_sas_url, filename, form_type, extracted_values, version_id)
    root = ET.Element("W2s")
    for extracted_value in extracted_values:
        w2_element = ET.SubElement(root, "W2")
        for key, value in extracted_value.items():
            if key == 'confidence':
                w2_element.set('confidence', str(value))
            else:
                ET.SubElement(w2_element, key).text = str(value) if value is not None else None
    xml_str = ET.tostring(root, encoding='utf-8').decode('utf-8')
    
    return ({"message": "W2 extraction successful", "xml": xml_str})

def sanitize_blob_name(blob_name):
    # Your existing sanitize function
    sanitized = re.sub(r'[()\[\]{}]', '', blob_name)
    sanitized = sanitized.replace(' ', '_')
    return sanitized

@app.route('/download_csv/<document_id>', methods=['GET', 'POST'])
async def download_csv(document_id):
    print("Entered download_csv function")
    json_data = await request.json
    client_id = json_data['clientID']
    db = Database(None, None)
    sanitized_doc_id = sanitize_blob_name(document_id)
    print(sanitized_doc_id)  
    csv_content = await db.generate_csv(sanitized_doc_id, client_id)

    await db.close()

    response = Response(csv_content, mimetype='text/csv')
    response.headers["Content-Disposition"] = f"attachment; filename={document_id}.csv" 
    
    return response

@app.route('/get_client_data', methods=['POST'])
async def get_client_data():
    json_data = await request.json
    client_id = json_data['clientID']

    # Use async with to ensure proper initialization and cleanup of TableBuilder
    async with TableBuilder() as table_builder:
        client_data = await table_builder.fetch_client_data(client_id)

    return jsonify({"data": client_data})

async def process_sort(file):
    filename = secure_filename(file.filename)
    print(f'Processing file: {filename}')
    file_stream = io.BytesIO(file.read())

    sorter = Sorter()
    result = await sorter.sort(file_stream)
    file_stream.close()

    return {**result, 'file_name': filename}

@app.route('/sort', methods=['POST'])
async def sort():
    files = await request.files
    files_to_sort = files.getlist('files[]')

    # Schedule all file processing tasks to run concurrently
    sorted_files = await asyncio.gather(*(process_sort(file) for file in files_to_sort))
    
    print(f'sorted_files: {sorted_files}')
    return {'sorted_files': sorted_files}

@app.route('/download_all_documents', methods=['POST', 'GET'])
async def download_all_documents():
        data = await request.json
        client_id = data['clientID']
        document_names = data['documentNames']

        # Initialize the BlobServiceClient asynchronously
        blob_service_client = BlobServiceClient.from_connection_string(azure_credentials.CONNECTION_STRING)
        container_client = blob_service_client.get_container_client(azure_credentials.BUCKET_NAME_CUSTOMS)

        # Download the existing FOFtest.xlsx
        blob_client = container_client.get_blob_client('FOFtemplate.xlsx')
        fof_test_stream = BytesIO()
        stream_downloader = await blob_client.download_blob()
        await stream_downloader.readinto(fof_test_stream)
        fof_test_stream.seek(0)
        fof_workbook = openpyxl.load_workbook(fof_test_stream)

        db = Database(None, None)

        # Create a new workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

        copy_worksheet(fof_workbook, workbook, 'Sheet1') 

        for document_name in document_names:
            sanitized_name = sanitize_blob_name(document_name)
            original_data, fof_data = await db.generate_sheet_data(sanitized_name, client_id)  
            
            # Add original sheet
            original_sheet = workbook.create_sheet(title=sanitized_name)
            for row in original_data:
                original_sheet.append(row)
            
            # Add FOF sheet
            fof_sheet = workbook.create_sheet(title=f"FOF_{sanitized_name}")
            for row in fof_data:
                fof_sheet.append(row)

        fof_sheet_objects = [workbook[sheet_name] for sheet_name in workbook.sheetnames if 'FOF_' in sheet_name]
        
        process_FOF(workbook, fof_sheet_objects) 

        # Save the workbook to a BytesIO object
        output_stream = BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)

        # Save the workbook to a temporary file
        async with aiofiles.tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            await tmp.write(output_stream.getvalue())
            await tmp.flush()
            response = await send_file(tmp.name, as_attachment=True, attachment_filename=f"{client_id}_Customs_Batch_Data.xlsx")
        return response



if __name__ == "__main__":
    app.run(debug=True)
    
    #main branch